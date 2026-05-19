from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from collections import Counter, OrderedDict
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

try:
    import openpyxl
except ImportError:  # pragma: no cover - exercised in environments without optional dependency
    openpyxl = None


ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = ROOT.parent

REQUIRED_HEADERS = {
    "SKU",
    "pid",
    "商家回复",
    "图片数量",
    "得分类型",
    "评论内容",
    "评论得分",
    "评论时间",
    "追评",
    "追评时间",
    "当前价格",
    "商品链接",
    "售后服务",
}

SHOP_NAME_HEADERS = ("店铺名称", "店铺名", "店铺名字", "shop_name")
PRODUCT_NAME_HEADERS = ("商品名称", "商品名", "产品名称", "product_name")
SAME_CELL_REFERENCE_RE = re.compile(r"^同\s*([A-Z]+[0-9]+)$", re.IGNORECASE)
JD_ITEM_ID_RE = re.compile(r"item\.jd\.com/(\d+)\.html", re.IGNORECASE)

XML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def _column_index(cell_ref: str) -> int:
    letters = re.sub(r"[^A-Z]", "", cell_ref.upper())
    index = 0
    for letter in letters:
        index = index * 26 + (ord(letter) - ord("A") + 1)
    return max(index - 1, 0)


def _shared_strings(zip_file: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zip_file.namelist():
        return []
    root = ET.fromstring(zip_file.read("xl/sharedStrings.xml"))
    values = []
    for item in root.findall("main:si", XML_NS):
        parts = [node.text or "" for node in item.findall(".//main:t", XML_NS)]
        values.append("".join(parts))
    return values


def _sheet_path(zip_file: zipfile.ZipFile, sheet_name: str) -> str:
    workbook = ET.fromstring(zip_file.read("xl/workbook.xml"))
    sheets = workbook.findall("main:sheets/main:sheet", XML_NS)
    if not sheets:
        raise SystemExit("Workbook has no sheets.")
    selected = next((sheet for sheet in sheets if sheet.get("name") == sheet_name), sheets[0])
    rel_id = selected.get(f"{{{XML_NS['rel']}}}id")
    rels = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))
    for rel in rels.findall("pkgrel:Relationship", XML_NS):
        if rel.get("Id") == rel_id:
            target = rel.get("Target", "")
            return f"xl/{target.lstrip('/')}" if not target.startswith("xl/") else target
    raise SystemExit(f"Cannot find worksheet relationship for sheet: {selected.get('name')}")


def _cell_value(cell, shared: list[str]):
    cell_type = cell.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//main:t", XML_NS))
    value_node = cell.find("main:v", XML_NS)
    if value_node is None:
        return ""
    raw = value_node.text or ""
    if cell_type == "s":
        try:
            return shared[int(raw)]
        except (ValueError, IndexError):
            return ""
    if cell_type == "b":
        return raw == "1"
    try:
        number = float(raw)
        return int(number) if number.is_integer() else number
    except ValueError:
        return raw


def _load_rows_with_stdlib(path: Path, sheet_name: str) -> list[list]:
    with zipfile.ZipFile(path) as zip_file:
        shared = _shared_strings(zip_file)
        worksheet_path = _sheet_path(zip_file, sheet_name)
        root = ET.fromstring(zip_file.read(worksheet_path))
        rows = []
        for row in root.findall("main:sheetData/main:row", XML_NS):
            values = []
            for cell in row.findall("main:c", XML_NS):
                index = _column_index(cell.get("r", "A1"))
                while len(values) <= index:
                    values.append("")
                values[index] = _cell_value(cell, shared)
            rows.append(values)
        return rows


def _load_workbook_rows(path: Path, sheet_name: str) -> list[list]:
    if openpyxl is None:
        return _load_rows_with_stdlib(path, sheet_name)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _text(value) -> str:
    return str(value).strip() if value is not None else ""


def _clean_text(value) -> str:
    return re.sub(r"\s+", " ", _text(value)).strip()


def _cell_by_reference(workbook_rows: list[list], cell_ref: str):
    row_number_text = re.sub(r"[^0-9]", "", cell_ref)
    if not row_number_text:
        return ""
    row_index = int(row_number_text) - 1
    col_index = _column_index(cell_ref)
    if row_index < 0 or row_index >= len(workbook_rows):
        return ""
    row = workbook_rows[row_index]
    return row[col_index] if col_index < len(row) else ""


def _resolve_same_cell_reference(value, workbook_rows: list[list], seen: set[str] | None = None) -> str:
    text = _clean_text(value)
    match = SAME_CELL_REFERENCE_RE.fullmatch(text)
    if not match:
        return text
    cell_ref = match.group(1).upper()
    seen = seen or set()
    if cell_ref in seen:
        return ""
    resolved = _resolve_same_cell_reference(_cell_by_reference(workbook_rows, cell_ref), workbook_rows, seen | {cell_ref})
    return resolved or text


def _is_same_cell_reference(value) -> bool:
    return bool(SAME_CELL_REFERENCE_RE.fullmatch(_clean_text(value)))


def _jd_item_id_from_url(url: str) -> str:
    match = JD_ITEM_ID_RE.search(_text(url))
    return match.group(1) if match else ""


def _to_float(value, default=None):
    if value in (None, ""):
        return default
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return default


def _to_int(value, default=0) -> int:
    numeric = _to_float(value)
    return int(numeric) if numeric is not None else default


def _strip_bought_prefix(text: str) -> str:
    text = _text(text)
    return text[2:].strip() if text.startswith("已购") else text


def _comment_type(rating, score_type: str) -> str:
    rating_value = _to_float(rating)
    if rating_value is not None:
        if rating_value >= 4.5:
            return "positive"
        if rating_value >= 3:
            return "neutral"
        return "negative"
    if any(word in score_type for word in ("差", "不满")):
        return "negative"
    if any(word in score_type for word in ("中", "一般")):
        return "neutral"
    return "positive"


def _capacity_from_titles(titles: list[str]) -> str:
    joined = " ".join(titles)
    values = [value for value in ("2-3人", "3-4人", "3-5人", "4-6人", "5-8人") if value in joined]
    return "/".join(values) if values else "multi-variant"


def _pick_price_groups(rows: list[dict], requested: list[float] | None) -> list[float]:
    if requested:
        return requested
    counts = Counter(row["price"] for row in rows if row.get("price") is not None)
    return sorted(counts)


def _representative_pid(items: list[dict]) -> str:
    counts = Counter(item["pid"] for item in items if item.get("pid"))
    return counts.most_common(1)[0][0] if counts else ""


def _product_group_key(row: dict) -> str:
    product_name = _clean_text(row.get("product_name") or row.get("shop_name"))
    if product_name:
        return f"name:{product_name}"
    item_id = _jd_item_id_from_url(row.get("product_url", ""))
    if item_id:
        return f"url:{item_id}"
    product_name = row.get("pid")
    return f"name:{_clean_text(product_name)}|price:{row.get('price')}"


def _optional_header(header_index: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for candidate in candidates:
        if candidate in header_index:
            return header_index[candidate]
    return None


def _most_common_text(items: list[dict], key: str, fallback: str = "") -> str:
    counts = Counter(item[key] for item in items if item.get(key))
    return counts.most_common(1)[0][0] if counts else fallback


def _after_sale_flags(service: str) -> dict:
    return {
        "free_shipping": "免费" in service or "京东发货" in service,
        "shipping_insurance": "退换" in service or "运费" in service,
        "return_7_days": "7天" in service or "七天" in service,
        "fast_refund": "闪电退款" in service or "极速审核" in service,
        "price_protection": "价保" in service,
        "official_store": False,
        "self_operated": "京东发货" in service,
    }


def _load_rows(path: Path, sheet_name: str) -> list[dict]:
    workbook_rows = _load_workbook_rows(path, sheet_name)
    if not workbook_rows:
        raise SystemExit("Workbook is empty.")
    headers = [_text(cell) for cell in workbook_rows[0]]
    header_index = {header: index for index, header in enumerate(headers) if header}
    missing = sorted(REQUIRED_HEADERS - set(header_index))
    if missing:
        raise SystemExit(f"Missing required headers: {', '.join(missing)}")
    shop_name_index = _optional_header(header_index, SHOP_NAME_HEADERS)
    product_name_index = _optional_header(header_index, PRODUCT_NAME_HEADERS)

    rows = []
    last_after_sale = ""
    last_shop_name = ""
    last_product_name = ""
    for raw in workbook_rows[1:]:
        price = _to_float(raw[header_index["当前价格"]])
        if price is None:
            continue
        raw_after_sale = _text(raw[header_index["售后服务"]])
        after_sale = _resolve_same_cell_reference(raw_after_sale, workbook_rows)
        if after_sale == _clean_text(raw_after_sale) and raw_after_sale.startswith("同") and last_after_sale:
            after_sale = last_after_sale
        if after_sale and not after_sale.startswith("同"):
            last_after_sale = after_sale
        shop_name = ""
        if shop_name_index is not None and len(raw) > shop_name_index:
            raw_shop_name = _text(raw[shop_name_index])
            shop_name = _resolve_same_cell_reference(raw_shop_name, workbook_rows)
            if _is_same_cell_reference(raw_shop_name) and shop_name == _clean_text(raw_shop_name) and last_shop_name:
                shop_name = last_shop_name
            if shop_name and not _is_same_cell_reference(shop_name):
                last_shop_name = shop_name
        product_name = ""
        if product_name_index is not None and len(raw) > product_name_index:
            raw_product_name = _text(raw[product_name_index])
            product_name = _resolve_same_cell_reference(raw_product_name, workbook_rows)
            if _is_same_cell_reference(raw_product_name) and product_name == _clean_text(raw_product_name) and last_product_name:
                product_name = last_product_name
            if product_name and not _is_same_cell_reference(product_name):
                last_product_name = product_name
        if not shop_name and product_name:
            shop_name = product_name
        rows.append(
            {
                "sku": _clean_text(raw[header_index["SKU"]]),
                "pid": _text(raw[header_index["pid"]]),
                "seller_reply": _clean_text(raw[header_index["商家回复"]]),
                "image_count": _to_int(raw[header_index["图片数量"]]),
                "score_type": _clean_text(raw[header_index["得分类型"]]),
                "comment_text": _clean_text(raw[header_index["评论内容"]]),
                "rating": _to_float(raw[header_index["评论得分"]]),
                "comment_time": _text(raw[header_index["评论时间"]]),
                "follow_up_text": _clean_text(raw[header_index["追评"]]),
                "follow_up_time": _text(raw[header_index["追评时间"]]),
                "price": price,
                "product_url": _text(raw[header_index["商品链接"]]),
                "after_sale": _clean_text(after_sale),
                "shop_name": shop_name,
                "product_name": product_name,
            }
        )
    return rows


def build_payload(input_path: Path, sheet_name: str, price_groups: list[float] | None = None) -> dict:
    rows = _load_rows(input_path, sheet_name)
    selected_prices = _pick_price_groups(rows, price_groups)
    selected_price_set = set(selected_prices)
    grouped: OrderedDict[str, list[dict]] = OrderedDict()
    for row in rows:
        if row["price"] in selected_price_set:
            grouped.setdefault(_product_group_key(row), []).append(row)

    canonical_products = []
    platform_products = []
    product_specs = []
    product_prices = []
    product_benefits = []
    return_policies = []
    comments = []
    summary = {}

    for index, (group_key, items) in enumerate(grouped.items(), start=1):
        price = items[0]["price"]
        representative_pid = _representative_pid(items)
        item_id = _jd_item_id_from_url(_most_common_text(items, "product_url"))
        external_group_id = f"jd-mvp-product-{item_id or representative_pid or index}"
        product_name = _most_common_text(items, "product_name")
        shop_name = _most_common_text(items, "shop_name")
        title = product_name or shop_name or f"京东帐篷价格组 {price:g}"
        titles = [item["sku"] for item in items if item["sku"]]
        pids = sorted({item["pid"] for item in items if item["pid"]})
        after_sale_texts = [item["after_sale"] for item in items if item["after_sale"]]
        after_sale = Counter(after_sale_texts).most_common(1)[0][0] if after_sale_texts else ""
        urls = [item["product_url"] for item in items if item["product_url"] and representative_pid in item["product_url"]]
        urls = urls or [item["product_url"] for item in items if item["product_url"]]
        positive_count = sum(1 for item in items if (item["rating"] or 0) >= 4)
        positive_rate = round(positive_count / len(items) * 100, 2) if items else 0.0

        canonical_products.append(
            {
                "external_group_id": external_group_id,
                "normalized_name": title,
                "brand": "",
                "model_name": title,
                "capacity": _capacity_from_titles(titles),
                "use_case": "",
                "main_image_url": None,
                "source": "jd_manual_xlsx_mvp",
            }
        )
        platform_products.append(
            {
                "external_group_id": external_group_id,
                "platform": "JD",
                "platform_product_id": representative_pid,
                "title": title,
                "shop_name": shop_name,
                "shop_type": "marketplace",
                "product_url": urls[0] if urls else f"https://item.jd.com/{representative_pid}.html",
                "image_url": None,
                "sales_volume": 0,
                "rating_count": len(items),
                "positive_rate": positive_rate,
            }
        )
        product_specs.append(
            {
                "platform_product_id": representative_pid,
                "setup_type": "quick_open_from_comment_summary",
                "tent_type": "family_or_weekend_camping",
                "raw_specs_json": {
                    "source_file": str(input_path),
                    "variant_pids": pids,
                    "top_sku_titles": [name for name, _count in Counter(titles).most_common(12)],
                    "source_product_name": product_name,
                    "source_shop_name": shop_name,
                    "note": "MVP keeps variant and package information as raw metadata; missing hard specs are not fabricated.",
                },
            }
        )
        product_prices.append(
            {
                "platform_product_id": representative_pid,
                "original_price": price,
                "current_price": price,
                "shop_coupon_amount": 0,
                "platform_coupon_amount": 0,
                "member_coupon_amount": 0,
                "limited_coupon_amount": 0,
                "red_packet_amount": 0,
                "discount_amount": 0,
                "shipping_fee": 0,
                "coupon_text": "No coupon field in source workbook; MVP uses visible current minimum price.",
                "promotion_text": "Current price is grouped by visible price in source workbook.",
                "price_update_time": datetime.now().date().isoformat(),
            }
        )
        product_benefits.append({"platform_product_id": representative_pid, **_after_sale_flags(after_sale), "gift_items": []})
        return_policies.append(
            {
                "platform_product_id": representative_pid,
                "return_shipping_insurance": "免费上门退换" in after_sale or "退换" in after_sale,
                "return_shipping_payer": "seller/platform if service text applies" if after_sale else "unknown",
                "return_condition_text": after_sale,
                "opened_return_allowed": "7天" in after_sale or "无理由" in after_sale,
                "used_return_allowed": False,
                "quality_issue_free_return": "免费上门退换" in after_sale or "假一赔四" in after_sale,
                "refund_speed_type": "fast_refund" if ("闪电退款" in after_sale or "极速审核" in after_sale) else "unknown",
                "refund_full_amount": True,
                "partial_refund_risk": False,
                "seller_return_attitude": "unknown",
                "return_policy_clarity": 75 if after_sale else 35,
            }
        )

        seen = set()
        skipped_duplicate = 0
        for item in items:
            if not item["comment_text"]:
                continue
            full_text = item["comment_text"]
            if item["follow_up_text"]:
                full_text = f'{full_text}\nFollow-up: {item["follow_up_text"]}'
            if item["sku"]:
                full_text = f"{full_text}\nPurchased variant: {_strip_bought_prefix(item['sku'])}"
            dedupe_key = re.sub(r"\s+", "", full_text)
            if dedupe_key in seen:
                skipped_duplicate += 1
                continue
            seen.add(dedupe_key)
            comments.append(
                {
                    "platform_product_id": representative_pid,
                    "platform": "JD",
                    "comment_text": full_text,
                    "rating": item["rating"],
                    "comment_type": _comment_type(item["rating"], item["score_type"]),
                    "has_image": item["image_count"] > 0,
                    "is_follow_up": bool(item["follow_up_text"]),
                    "comment_time": item["follow_up_time"] if item["follow_up_text"] and item["follow_up_time"] else item["comment_time"],
                    "seller_reply": item["seller_reply"] or None,
                }
            )
        summary[representative_pid] = {
            "product_group_key": group_key,
            "price_group": price,
            "source_rows": len(items),
            "deduped_comments": len(seen),
            "skipped_duplicate": skipped_duplicate,
            "variant_pid_count": len(pids),
            "representative_title": title,
        }

    return {
        "source_name": "jd_manual_xlsx_mvp",
        "source_file": str(input_path),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "_warnings": [
            "MVP groups all real JD SKU/package variants into main products by visible price group.",
            "Detailed waterproof/weight/size specs are missing in source workbook and are not fabricated.",
            "Nickname, comment id, image URLs and video URLs are excluded from import payload.",
        ],
        "_cleaning_summary": summary,
        "canonical_products": canonical_products,
        "platform_products": platform_products,
        "product_specs": product_specs,
        "product_prices": product_prices,
        "product_benefits": product_benefits,
        "return_policies": return_policies,
        "comments": comments,
        "redbook_notes": [],
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Build a CampRank MVP JSON payload from a JD comment workbook.")
    parser.add_argument("--input", required=True, help="Path to the JD workbook, for example ../data.xlsx")
    parser.add_argument("--output", default="data/real_samples/jd_tents_mvp.json")
    parser.add_argument("--sheet", default="data")
    parser.add_argument("--price-groups", nargs="*", type=float, default=None)
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.is_absolute():
        input_path = (Path.cwd() / input_path).resolve()
    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = (ROOT / output_path).resolve()

    payload = build_payload(input_path, args.sheet, args.price_groups)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"output_path": str(output_path), "comments": len(payload["comments"]), "summary": payload["_cleaning_summary"]}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
